from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
import os

# Setup Jinja2 environment
env = Environment(loader=FileSystemLoader('templates'))
template = env.get_template('config_from_xls.j2')

# Load Excel workbook
wb = load_workbook(filename='var_table.xlsx')
ws = wb['bgp']

# Get headers (variable names)
headers = [cell.value for cell in ws[1]]

# Process each row (host)
for row in range(2, ws.max_row + 1):
    # Create variables dictionary for this host
    host_vars = {}
    
    # Process each column
    for col in range(1, ws.max_column + 1):
        var_name = headers[col-1]
        var_value = ws.cell(row=row, column=col).value
        host_vars[var_name] = var_value
    
    # Render template
    rendered_config = template.render(host_vars)
    
    # Save to file
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    hostname = host_vars.get('hostname', f'device_{row}')
    with open(f'{output_dir}/{hostname}.conf', 'w') as f:
        f.write(rendered_config)
    
    print(f"Configuration for {hostname} rendered to {output_dir}/{hostname}.conf")

